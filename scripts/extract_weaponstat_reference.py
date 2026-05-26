#!/usr/bin/env python3
"""Extract PvP damage reference rows from the shared WeaponStat Google Sheet."""

from __future__ import annotations

import csv
import io
import math
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SHEET_ID = "1FWMC-Vd_bGEoRkkrn3drWIORCFYyWQVMNlMasa4OE6I"
SOURCE_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
SOURCE_SHEET = "WeaponStat Chart"
OUTPUT = ROOT / "data" / "static" / "ttk" / "weaponstat_community_reference.csv"
TARGET_HP = 230

AMMO_CODES = {
    "Primary": "p",
    "Special": "s",
    "Heavy": "h",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value).strip()


def number(value: Any) -> float | None:
    raw = clean(value)
    if not raw or raw in {".", "?", "[none]"}:
        return None
    try:
        return float(raw.rstrip("%"))
    except ValueError:
        return None


def ms_from_seconds(value: Any) -> str:
    seconds = number(value)
    if seconds is None:
        return ""
    return clean(round(seconds * 1000, 3))


def shots_to_kill(damage: float | None) -> str:
    if damage is None or damage <= 0:
        return ""
    return str(math.ceil(TARGET_HP / damage))


def parse_optimal_mix(value: Any) -> tuple[str, str]:
    text = clean(value).lower()
    if not text or text in {".", "[none]"}:
        return "", ""
    heads = ""
    bodies = ""
    if "h" in text:
        heads = text.split("h", 1)[0]
        rest = text.split("h", 1)[1]
        if "b" in rest:
            bodies = rest.split("b", 1)[0]
    elif "b" in text:
        bodies = text.split("b", 1)[0]
    return heads, bodies


def body_ttk_ms(row: tuple[Any, ...], body_shots: str, optimal_ttk: str, optimal_mix: str) -> str:
    if not body_shots:
        return ""
    try:
        shots = int(float(body_shots))
    except ValueError:
        return ""
    if shots <= 1:
        return "0"

    _, optimal_bodies = parse_optimal_mix(optimal_mix)
    if optimal_bodies and not clean(optimal_mix).lower().split("b", 1)[-1]:
        try:
            if int(float(optimal_bodies)) == shots and optimal_ttk:
                return optimal_ttk
        except ValueError:
            pass

    rpm = number(row[4] if len(row) > 4 else None)
    if rpm and rpm > 0:
        return clean(round((shots - 1) * 60000 / rpm, 3))
    return ""


def extract_records(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    worksheet = workbook[SOURCE_SHEET]
    records: list[dict[str, Any]] = []
    sandbox_version = clean(worksheet.cell(row=5, column=2).value) or "WeaponStat community sheet"

    for row_number, row in enumerate(worksheet.iter_rows(min_row=8, values_only=True), start=8):
        weapon_type = clean(row[0] if len(row) > 0 else "")
        archetype = clean(row[1] if len(row) > 1 else "")
        ammo_label = clean(row[5] if len(row) > 5 else "")
        lookup = clean(row[7] if len(row) > 7 else "")
        if not weapon_type or not archetype or ammo_label not in AMMO_CODES or not lookup:
            continue

        body_damage = number(row[2] if len(row) > 2 else None)
        crit_damage = number(row[3] if len(row) > 3 else None)
        if body_damage is None and crit_damage is None:
            continue
        if body_damage == 0 and crit_damage == 0:
            continue

        body_shots = shots_to_kill(body_damage)
        crit_shots = shots_to_kill(crit_damage)
        optimal_ttk = ms_from_seconds(row[21] if len(row) > 21 else None)
        optimal_mix = clean(row[20] if len(row) > 20 else "")
        optimal_heads, optimal_bodies = parse_optimal_mix(optimal_mix)
        optimal_total = clean(row[19] if len(row) > 19 else "")

        forgiveness_count = optimal_bodies
        forgiveness_pct = ""
        if forgiveness_count and optimal_total:
            try:
                total = float(optimal_total)
                count = float(forgiveness_count)
                if total > 0:
                    forgiveness_pct = clean(round(count / total, 4))
            except ValueError:
                forgiveness_pct = ""

        note = " / ".join(part for part in (clean(row[14] if len(row) > 14 else ""), clean(row[15] if len(row) > 15 else ""), clean(row[16] if len(row) > 16 else "")) if part)
        records.append(
            {
                "source_url": SOURCE_URL,
                "source_sheet": SOURCE_SHEET,
                "source_row": row_number,
                "sandbox_version": sandbox_version,
                "lookup": lookup,
                "sort": clean(row[6] if len(row) > 6 else ""),
                "weapon_family": weapon_type,
                "ammo_code": AMMO_CODES[ammo_label],
                "ammo_label": ammo_label,
                "archetype": archetype,
                "note": note,
                "body_damage": clean(body_damage),
                "crit_damage": clean(crit_damage),
                "rpm": clean(row[4] if len(row) > 4 else ""),
                "shot_delay": clean(row[8] if len(row) > 8 else ""),
                "burst_delay": clean(row[9] if len(row) > 9 else ""),
                "sub_burst_delay": clean(row[10] if len(row) > 10 else ""),
                "charge_ms": clean(row[11] if len(row) > 11 else ""),
                "shots_per_burst": clean(row[12] if len(row) > 12 else ""),
                "shots_per_sub": clean(row[13] if len(row) > 13 else ""),
                "crit_shots": crit_shots,
                "body_shots": body_shots,
                "optimal_kill": optimal_mix,
                "optimal_shots": optimal_total,
                "optimal_ttk_ms": optimal_ttk,
                "body_ttk_ms": body_ttk_ms(row, body_shots, optimal_ttk, optimal_mix),
                "body_forgiveness_count": forgiveness_count,
                "body_forgiveness_pct": forgiveness_pct,
                "reference_status": "WeaponStat community reference",
                "verification_policy": "Prefer over local DrYamaHiro sheet; re-check against Bungie/API when official numeric changes exist.",
            }
        )
    return records


def main() -> None:
    with urllib.request.urlopen(EXPORT_URL, timeout=60) as response:
        workbook_bytes = response.read()
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    records = extract_records(workbook)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0].keys()) if records else [])
        writer.writeheader()
        writer.writerows(records)
    print({"source": SOURCE_URL, "output": str(OUTPUT), "records": len(records)})


if __name__ == "__main__":
    main()
