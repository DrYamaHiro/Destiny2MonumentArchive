from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "fromDrYamaHiro" / "D2WP_boost_ver1.02.xlsx"
OUTPUT = ROOT / "data" / "static" / "ttk" / "dr_yamahiro_wp_reference.csv"
REFERENCE_SHEETS = ("PrimaryWeapons", "SpecialWeapons", "GutshotStraight")


def clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
        if text in {"‚È‚µ", "なし"}:
            return "なし"
        return text
    return value


def number_ms(value: Any) -> Any:
    value = clean(value)
    if isinstance(value, (int, float)):
        return round(value * 1000, 3)
    return value


def rows_from_sheet(workbook: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    worksheet = workbook[sheet_name]
    rows: list[dict[str, Any]] = []
    weapon_family = ""
    for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None and cell != "" for cell in row):
            continue
        if clean(row[0]):
            weapon_family = clean(row[0])
        archetype = clean(row[1])
        if not weapon_family or not archetype:
            continue
        rows.append(
            {
                "source_workbook": SOURCE.name,
                "source_sheet": sheet_name,
                "source_row": index,
                "weapon_family": weapon_family,
                "archetype": archetype,
                "burst_count": clean(row[2]) if len(row) > 2 else "",
                "burst_delay_sec": clean(row[3]) if len(row) > 3 else "",
                "rpm": clean(row[4]) if len(row) > 4 else "",
                "crit_damage": clean(row[5]) if len(row) > 5 else "",
                "body_damage": clean(row[6]) if len(row) > 6 else "",
                "optimal_ttk_ms": number_ms(row[7]) if len(row) > 7 else "",
                "crit_shots": clean(row[8]) if len(row) > 8 else "",
                "body_ttk_ms": number_ms(row[9]) if len(row) > 9 else "",
                "body_shots": clean(row[10]) if len(row) > 10 else "",
                "wp_to_boost_hsttk": clean(row[11]) if len(row) > 11 else "",
                "reference_status": "DrYamaHiro reference only",
                "verification_policy": "Do not apply to PvP Potential until checked against Bungie/API/patch-note primary sources.",
            }
        )
    return rows


def main() -> None:
    workbook = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    records: list[dict[str, Any]] = []
    for sheet_name in REFERENCE_SHEETS:
        if sheet_name in workbook.sheetnames:
            records.extend(rows_from_sheet(workbook, sheet_name))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0].keys()) if records else [])
        writer.writeheader()
        writer.writerows(records)
    print({"source": str(SOURCE), "output": str(OUTPUT), "records": len(records)})


if __name__ == "__main__":
    main()
